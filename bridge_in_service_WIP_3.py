from datetime import datetime, timedelta
import calendar
import openpyxl
import win32com.client as win32
import textwrap
import os
from dateutil.relativedelta import relativedelta


class DateOperations:
    test_date = None
    @staticmethod
    def get_todays_date():
        if DateOperations.test_date is not None:
            return DateOperations.test_date
        return datetime.now()
    
    @staticmethod
    def reset_test_date():
        DateOperations.test_date = None

    @staticmethod
    def set_test_date(date):
        DateOperations.test_date = date

    @staticmethod
    def convert_to_datetime(date_str, date_format="%m/%d/%Y"):
        try:
            return datetime.strptime(date_str, date_format)
        except ValueError:
            return None


class UserInput:

    @staticmethod
    def ask_for_employee_id():
        return input("Please enter the employee's ID: ")

    @staticmethod
    def ask_for_first_name():
        return input("Please enter the employee's first name: ")

    @staticmethod
    def ask_for_last_name():

        return input("Please enter the employee's last name: ")

    @staticmethod
    def ask_for_most_recent_start_date():

        return input("Please enter the most recent start date (MM/DD/YYYY): ")

    @staticmethod
    def ask_for_employment_period_start():

        return input("Enter the start date of the employment period (MM/DD/YYYY): ")

    @staticmethod
    def ask_for_employment_period_end():

        return input("Enter the end date of the employment period (MM/DD/YYYY): ")

    @staticmethod
    def ask_for_employee_fte():

        return input("Please enter the employee's FTE as of their most recent start date (between 0.75 and 1.0): ")

    @staticmethod
    def ask_for_date_of_fte_change():

        return input("Enter the date of the FTE change (MM/DD/YYYY): ")

    @staticmethod
    def ask_for_new_fte():

        return input("Enter the new FTE (between 0.75 and 1.0): ")

    @staticmethod
    def ask_yes_no_or_quit(prompt):

        response = input(prompt).strip().lower()
        while response not in ['yes', 'no', 'quit']:
            print("Invalid input. Please enter 'yes', 'no', or 'quit'.")
            response = input(prompt).strip().lower()
        return response

    @staticmethod
    def ask_to_add_another_period():

        return UserInput.ask_yes_no_or_quit("Do you want to add another employment period? (yes/no/quit): ")

    @staticmethod
    def ask_if_fte_changed():

        return UserInput.ask_yes_no_or_quit("Did the FTE change after the most recent start date? (yes/no/quit): ")


class Employee:

    def __init__(self, employee_id, first_name, last_name, most_recent_start_date, initial_fte):

        self.employee_id = employee_id
        self.first_name = first_name
        self.last_name = last_name
        self.most_recent_start_date = most_recent_start_date
        self.prior_employment_periods = [(most_recent_start_date, DateOperations.get_todays_date())]
        self.fte_changes = [(most_recent_start_date, initial_fte)]
        self.bridge_in_service_date = None
        self.pto_accrual_difference = 0
        # Initialize accrual data attributes
        self.original_monthly_accruals = {}
        self.bridge_monthly_accruals = {}
        self.accrual_differences = {}


    def get_employee_id(self):

        return self.employee_id

    def get_first_name(self):

        return self.first_name

    def get_last_name(self):

        return self.last_name

    def get_most_recent_start_date(self):

        return self.most_recent_start_date

    def get_fte(self):
 
        return self.fte

    def add_employment_period(self, start_date, end_date):

        self.prior_employment_periods.append((start_date, end_date))

    def get_employment_periods(self):

        return self.prior_employment_periods

    def add_fte_change(self, change_date, new_fte):

        self.fte_changes.append((change_date, new_fte))

    def get_fte_changes(self):

        return self.fte_changes

    def set_bridge_in_service_date(self, bridge_date):

        self.bridge_in_service_date = bridge_date

    def update_pto_accrual_difference(self, difference):

        self.pto_accrual_difference = difference


class Verification:

    @staticmethod
    def verify_employee_id(employee_id):

        if len(employee_id) == 8 and employee_id.isdigit():
            return True, "Valid ID."
        else:
            return False, "Invalid ID. Please ensure it is 8 digits."

    @staticmethod
    def verify_most_recent_start_date(start_date_str):

        start_date = DateOperations.convert_to_datetime(start_date_str)
        if start_date is None:
            return False, "Invalid date format. Please use MM/DD/YYYY."

        today = DateOperations.get_todays_date()
        six_months_ago = today - timedelta(days=182)  # More accurate approximation

        if start_date > today:
            return False, "Start date cannot be in the future."
        elif start_date > six_months_ago:
            return False, "Start date must be at least 6 months ago."
        else:
            return True, "Valid start date."
    
    @staticmethod
    def verify_date(date_str):

        date = DateOperations.convert_to_datetime(date_str)
        if date is None:
            return False, "Invalid date format. Please use MM/DD/YYYY."
        today = DateOperations.get_todays_date()
        if date > today:
            return False, "Date must not be in the future."
        return True, "Valid date."

    @staticmethod
    def verify_no_overlap(new_start_date, new_end_date, existing_periods):

        for existing_start, existing_end in existing_periods:
            # Check if new period starts or ends within an existing period
            if (new_start_date <= existing_end and new_end_date >= existing_start) or \
               (new_end_date >= existing_start and new_end_date <= existing_end):
                return False, "Period overlaps with an existing period."
        return True, "No overlap with existing periods."

    @staticmethod
    def verify_employee_fte(fte):

        try:
            fte_value = float(fte)
        except ValueError:
            return False, "Invalid FTE value. FTE must be a number."

        if 0.75 <= fte_value <= 1.0:
            return True, "Valid FTE."
        else:
            return False, "FTE must be between 0.75 and 1.0."

    @staticmethod
    def verify_employment_period_start(start_date, existing_periods, most_recent_start_date):

        if start_date >= most_recent_start_date:
            return False, "Start date must be before the most recent start date."

        for existing_start, existing_end in existing_periods:
            if start_date >= existing_start and start_date <= existing_end:
                return False, "Start date overlaps with an existing period."
        return True, "Start date is valid."

    @staticmethod
    def verify_no_overlap_with_end_date(new_start_date, new_end_date, existing_periods):

        for start_date, end_date in existing_periods:
            # Overlap occurs if the new period starts before an existing period ends AND the new period ends after the existing period starts
            if new_start_date <= end_date and new_end_date >= start_date:
                return False, "Period overlaps with an existing period."
        return True, "End date is valid."


class Calculation:

    @staticmethod
    def calculate_service_months_from_recent_start(employee):
        # Calculate the bridge in service date
        today = DateOperations.get_todays_date()
        most_recent_start_date = employee.most_recent_start_date  

        # Adjust the start date to the next 16th after the bridge in service date
        if most_recent_start_date.day >= 16:
            start_date = most_recent_start_date.replace(day=16) + timedelta(days=32)
            start_date = start_date.replace(day=16)
        else:
            start_date = most_recent_start_date.replace(day=16)


        # Get today's date
        today = DateOperations.get_todays_date()

        # Calculate full months between the adjusted start date and today
        if today > start_date:
            total_months = (today.year - start_date.year) * 12 + (today.month - start_date.month)
        else:
            total_months = 0
        
        if today.day < 16:
            total_months += 1
            if most_recent_start_date.day < 16:
                total_months -= 1
        
        return total_months
    
    @staticmethod
    def calculate_service_months_from_recent_start_pre_16(employee):
        # Get the most recent start date
        today = DateOperations.get_todays_date()
        most_recent_start_date = employee.most_recent_start_date

        # Determine the end of the month for the current month
        end_of_current_month = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])

        start_date = most_recent_start_date
        end_date = end_of_current_month

        # Calculate the full months between the start date and the end date
        month_count = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
        
        # If the end date's day is greater than or equal to the start date's day, add one more month
        if end_date.day >= start_date.day:
            month_count += 1
            
        return month_count

    @staticmethod
    def calculate_service_months_from_bridge(employee):
        # Calculate the bridge in service date
        bridge_in_service_date = employee.bridge_in_service_date

        # Adjust the start date to the next 16th after the bridge in service date
        if bridge_in_service_date.day >= 16:
            start_date = bridge_in_service_date.replace(day=16) + timedelta(days=32)
            start_date = start_date.replace(day=16)
        else:
            start_date = bridge_in_service_date.replace(day=16)


        # Get today's date
        today = DateOperations.get_todays_date()

        # Calculate full months between the adjusted start date and today
        if today > start_date:
            total_months = (today.year - start_date.year) * 12 + (today.month - start_date.month)
        else:
            total_months = 0
        
        return total_months
    
    @staticmethod
    def calculate_service_months_from_bridge_pre_16(employee):
        # Get today's date
        today = DateOperations.get_todays_date()

        # Get the bridge in service date
        bridge_in_service_date = employee.bridge_in_service_date

        # Determine the end of the current month for any given day
        end_of_current_month = datetime(today.year, today.month, calendar.monthrange(today.year, today.month)[1])

        start_date = bridge_in_service_date
        end_date = end_of_current_month

        # Calculate the full months between the start date and the end date
        month_count = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
        
        # If the end date's day is greater than or equal to the start date's day, add one more month
        if end_date.day >= start_date.day:
            month_count += 1
            
        return month_count
    
    

    @staticmethod
    def get_accrual_rate_for_months_of_service(months_of_service, fte, employee):
        """
        Calculates the PTO accrual rate based on the number of months of service and full-time equivalence (FTE).

        Parameters:
        months_of_service (int): The total months of service.
        fte (float): The full-time equivalence factor.

        Returns:
        float: The calculated accrual rate per month adjusted by FTE.
        """
        # Define the accrual rates
        thresholds = [
            (60, 13.33),
            (120, 16.66),
            (float('inf'), 20)  # This effectively handles any case above 120 months
        ]

        # Determine the accrual rate based on the thresholds
        for threshold, rate in thresholds:
            if employee.bridge_in_service_date.day >= 16:
                if months_of_service < threshold:
                    return rate * fte
            if months_of_service <= threshold:
                return rate * fte

        # Should not reach here due to the last threshold of float('inf'), but just in case:
        return thresholds[-1][1] * fte


    @staticmethod
    def calculate_pto_accrual_rate(employee):
        total_service_months = Calculation.calculate_service_months_from_recent_start(employee)
        total_service_months_pre_16 = Calculation.calculate_service_months_from_recent_start_pre_16(employee)
        total_pto_accrued = 0
        accrual_details = {}  # Dictionary to store detailed accruals per month
        today = DateOperations.get_todays_date()

        # Compute each month's effective start and end date, then calculate accrual
        if today.day < 16:
            for i in range(1, total_service_months):
                month_incremented = (employee.most_recent_start_date.month + i - 1) % 12 + 1
                year_incremented = employee.most_recent_start_date.year + (employee.most_recent_start_date.month + i - 1) // 12

                effective_month_start = datetime(year_incremented, month_incremented, 1)
                days_in_month = calendar.monthrange(year_incremented, month_incremented)[1]
                effective_month_end = datetime(year_incremented, month_incremented, days_in_month)

                current_fte = Calculation.update_fte_based_on_changes(employee, effective_month_start, effective_month_end)
                adjusted_service_months = Calculation.calculate_adjusted_service_months_for_most_recent(total_service_months, total_service_months, i, employee)
                pto_accrued_this_month = Calculation.get_accrual_rate_for_months_of_service(adjusted_service_months, current_fte, employee)

                month_year = effective_month_start.strftime("%B %Y")
                accrual_details[month_year] = f"{pto_accrued_this_month:.2f} Hours (FTE: {current_fte:.2f})"

                total_pto_accrued += pto_accrued_this_month
        else:
            for i in range(1, total_service_months_pre_16):
                month_incremented = (employee.most_recent_start_date.month + i - 1) % 12 + 1
                year_incremented = employee.most_recent_start_date.year + (employee.most_recent_start_date.month + i - 1) // 12

                effective_month_start = datetime(year_incremented, month_incremented, 1)
                days_in_month = calendar.monthrange(year_incremented, month_incremented)[1]
                effective_month_end = datetime(year_incremented, month_incremented, days_in_month)

                current_fte = Calculation.update_fte_based_on_changes(employee, effective_month_start, effective_month_end)
                adjusted_service_months = Calculation.calculate_adjusted_service_months_for_most_recent_post_16(total_service_months_pre_16, total_service_months_pre_16, i, employee)
                pto_accrued_this_month = Calculation.get_accrual_rate_for_months_of_service(adjusted_service_months, current_fte, employee)

                month_year = effective_month_start.strftime("%B %Y")
                accrual_details[month_year] = f"{pto_accrued_this_month:.2f} Hours (FTE: {current_fte:.2f})"

                total_pto_accrued += pto_accrued_this_month

        return round(total_pto_accrued, 2), accrual_details
    
    @staticmethod
    def calculate_accrual_differences(accruals_dict1, accruals_dict2):
        """
        Calculate the differences between two accrual dictionaries, and sort by month.

        Parameters:
        accruals_dict1 (dict): The first dictionary of accruals.
        accruals_dict2 (dict): The second dictionary of accruals.

        Returns:
        dict: A dictionary with the differences in accruals, formatted as strings and sorted by month.
        """
        accrual_differences = {}
        all_months = set(accruals_dict1.keys()).union(accruals_dict2.keys())

        # Convert month names to datetime to sort them
        sorted_months = sorted(all_months, key=lambda date: datetime.strptime(date, "%B %Y"))

        for month in sorted_months:
            accrual1 = float(accruals_dict1.get(month, "0.00 Hours").split()[0])
            accrual2 = float(accruals_dict2.get(month, "0.00 Hours").split()[0])
            difference = accrual2 - accrual1
            accrual_differences[month] = f"{difference:.2f} Hours"

        return accrual_differences
    
    @staticmethod
    def calculate_adjusted_service_months_for_most_recent(total_service_months, service_months_since_recent_start, month_index, employee):
        """
        Calculate the adjusted service months taking into account the overall service time and the specific month offset.

        Parameters:
        total_service_months (int): Total number of service months calculated from the start date.
        service_months_since_recent_start (int): Service months calculated from the most recent start date.
        month_index (int): Current index in the loop to adjust for the right month.

        Returns:
        int: Adjusted service months for the specific month.
        """
        if employee.most_recent_start_date.day >= 16:
            # Include the current month in the accrual
            total_service_months -= 1
        
        return total_service_months - service_months_since_recent_start + month_index
    
    @staticmethod
    def calculate_adjusted_service_months_for_most_recent_post_16(total_service_months, service_months_since_recent_start, month_index, employee):
        """
        Calculate the adjusted service months taking into account the overall service time and the specific month offset.

        Parameters:
        total_service_months (int): Total number of service months calculated from the start date.
        service_months_since_recent_start (int): Service months calculated from the most recent start date.
        month_index (int): Current index in the loop to adjust for the right month.

        Returns:
        int: Adjusted service months for the specific month.
        """
        if employee.most_recent_start_date.day > 15 :
            total_service_months -= 1
        
        return total_service_months - service_months_since_recent_start + month_index
    
    @staticmethod
    def calculate_adjusted_service_months_for_bridge(total_service_months, service_months_since_recent_start, month_index, employee):
        """
        Calculate the adjusted service months taking into account the overall service time and the specific month offset.

        Parameters:
        total_service_months (int): Total number of service months calculated from the start date.
        service_months_since_recent_start (int): Service months calculated from the most recent start date.
        month_index (int): Current index in the loop to adjust for the right month.

        Returns:
        int: Adjusted service months for the specific month.
        """
        if employee.most_recent_start_date.day < 16 :
            total_service_months += 1
        if employee.bridge_in_service_date.day < 16:
            total_service_months -= 1
            if employee.most_recent_start_date.day > 15:
                total_service_months += 1
        if employee.bridge_in_service_date.day > 15:
            # Include the current month in the accrual
            total_service_months -= 1
            if employee.most_recent_start_date.day > 15:
                total_service_months += 1
        
        return total_service_months - service_months_since_recent_start + month_index
    
    @staticmethod
    def calculate_adjusted_service_months_for_bridge_post_16(total_service_months, service_months_since_recent_start, month_index, employee):
        """
        Calculate the adjusted service months taking into account the overall service time and the specific month offset.

        Parameters:
        total_service_months (int): Total number of service months calculated from the start date.
        service_months_since_recent_start (int): Service months calculated from the most recent start date.
        month_index (int): Current index in the loop to adjust for the right month.

        Returns:
        int: Adjusted service months for the specific month.
        """
        if employee.most_recent_start_date.day > 15 :
            total_service_months -= 1
            if employee.bridge_in_service_date.day > 15:
                total_service_months -= 1
        if employee.most_recent_start_date.day < 16:
            total_service_months -= 1
            if employee.bridge_in_service_date.day > 15:
                total_service_months -= 1
        if employee.bridge_in_service_date.day > 15:
            # Include the current month in the accrual
            total_service_months += 1
        if employee.bridge_in_service_date.day < 16:
            total_service_months += 1
        return total_service_months - service_months_since_recent_start + month_index
    
    def update_fte_based_on_changes(employee, current_month_start, current_month_end):
        """
        Update the FTE based on changes that fall within the specified month period.
        Once an FTE change is applicable, it should continue to be applied forward until a new change overrides it.

        Parameters:
        employee (Employee): The employee whose FTE changes are to be updated.
        current_month_start (datetime): The start of the current month period.
        current_month_end (datetime): The end of the current month period.

        Returns:
        float: Updated FTE after applying relevant changes.
        """
        current_fte = employee.fte_changes[0][1]  # Start with the initial FTE set at hiring
        applicable_fte = current_fte  # This will store the most recently applicable FTE change

        for change_date, new_fte in sorted(employee.fte_changes, key=lambda x: x[0]):
            # Check if the change is within the current month and before the 15th or in previous months
            if change_date <= current_month_end:
                if change_date.day <= 15 or change_date < current_month_start:
                    applicable_fte = new_fte
                elif change_date.day > 15 and change_date.month == current_month_start.month:
                    # Changes after the 15th should apply from the start of the next month
                    next_month_start = current_month_end.replace(day=1) + timedelta(days=32)
                    next_month_start = next_month_start.replace(day=1)
                    if current_month_start >= next_month_start:
                        applicable_fte = new_fte

        return applicable_fte

    @staticmethod
    def calculate_bridge_pto_accrual_rate(employee):
        total_service_months = Calculation.calculate_service_months_from_bridge(employee)
        total_service_months_pre_16 = Calculation.calculate_service_months_from_bridge_pre_16(employee)
        service_months_since_recent_start = Calculation.calculate_service_months_from_recent_start(employee)
        service_months_since_recent_start_pre_16 = Calculation.calculate_service_months_from_recent_start_pre_16(employee)
        today = DateOperations.get_todays_date()

        total_pto_accrued = 0
        accrual_details = {}  # Dictionary to store detailed accruals per month
        if today.day < 16:
            for i in range(1, service_months_since_recent_start):  # Start from 1 to skip the first month
                month_incremented = (employee.most_recent_start_date.month + i - 1) % 12 + 1
                year_incremented = employee.most_recent_start_date.year + (employee.most_recent_start_date.month + i - 1) // 12

                effective_month_start = datetime(year_incremented, month_incremented, 1)
                days_in_month = calendar.monthrange(year_incremented, month_incremented)[1]
                effective_month_end = datetime(year_incremented, month_incremented, days_in_month)

                current_fte = Calculation.update_fte_based_on_changes(employee, effective_month_start, effective_month_end)
                adjusted_service_months = Calculation.calculate_adjusted_service_months_for_bridge(total_service_months, service_months_since_recent_start, i, employee)
                pto_accrued_this_month = Calculation.get_accrual_rate_for_months_of_service(adjusted_service_months, current_fte, employee)

                month_year = effective_month_start.strftime("%B %Y")
                accrual_details[month_year] = f"{pto_accrued_this_month:.2f} Hours (FTE: {current_fte:.2f})"

                total_pto_accrued += pto_accrued_this_month
        else:
            for i in range(1, service_months_since_recent_start_pre_16):  # Start from 1 to skip the first month
                month_incremented = (employee.most_recent_start_date.month + i - 1) % 12 + 1
                year_incremented = employee.most_recent_start_date.year + (employee.most_recent_start_date.month + i - 1) // 12

                effective_month_start = datetime(year_incremented, month_incremented, 1)
                days_in_month = calendar.monthrange(year_incremented, month_incremented)[1]
                effective_month_end = datetime(year_incremented, month_incremented, days_in_month)

                current_fte = Calculation.update_fte_based_on_changes(employee, effective_month_start, effective_month_end)
                adjusted_service_months = Calculation.calculate_adjusted_service_months_for_bridge_post_16(total_service_months_pre_16, service_months_since_recent_start_pre_16, i, employee)
                pto_accrued_this_month = Calculation.get_accrual_rate_for_months_of_service(adjusted_service_months, current_fte, employee)

                month_year = effective_month_start.strftime("%B %Y")
                accrual_details[month_year] = f"{pto_accrued_this_month:.2f} Hours (FTE: {current_fte:.2f})"

                total_pto_accrued += pto_accrued_this_month
        return round(total_pto_accrued, 2), accrual_details

    @staticmethod
    def calculate_total_service_duration(prior_employment_periods):

        total_duration = timedelta(days=0)
        for start_date, end_date in prior_employment_periods:
            period_duration = end_date - start_date
            total_duration += period_duration
        return total_duration

    @staticmethod
    def calculate_bridge_in_service_date(employee):

        total_service_duration = Calculation.calculate_total_service_duration(employee.prior_employment_periods)
        bridge_in_service_date = DateOperations.get_todays_date() - total_service_duration
        employee.set_bridge_in_service_date(bridge_in_service_date)
        return bridge_in_service_date
    
class ExcelExport:

    @staticmethod
    def try_export_employee_data(employee, directory_path, original_monthly_accruals, bridge_monthly_accruals, accrual_differences):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            # Populate bridge in service data
            info_titles = ["Employee ID", "First Name", "Last Name", "Most Recent Start Date", "Bridge In Service Date", "PTO Accrual Difference"]
            info_values = [
                employee.employee_id,
                employee.first_name,
                employee.last_name,
                employee.most_recent_start_date.strftime("%m/%d/%Y"),
                employee.bridge_in_service_date.strftime("%m/%d/%Y") if employee.bridge_in_service_date else "N/A",
                employee.pto_accrual_difference
            ]

            for i, title in enumerate(info_titles, start=1):
                ws[f'A{i}'] = title
                ws[f'B{i}'] = info_values[i-1]

            # Calculate start column for accrual data
            accrual_start_col = 'D'
            ws[f'{accrual_start_col}1'] = "Month Year"
            ws[f'{chr(ord(accrual_start_col)+1)}1'] = "Original"
            ws[f'{chr(ord(accrual_start_col)+2)}1'] = "Bridge"
            ws[f'{chr(ord(accrual_start_col)+3)}1'] = "Difference"
            accrual_row = 2

            # Add accrual data rows
            for month in accrual_differences:
                ws[f'{accrual_start_col}{accrual_row}'] = month
                ws[f'{chr(ord(accrual_start_col)+1)}{accrual_row}'] = original_monthly_accruals.get(month, '0.00 Hours')
                ws[f'{chr(ord(accrual_start_col)+2)}{accrual_row}'] = bridge_monthly_accruals.get(month, '0.00 Hours')
                ws[f'{chr(ord(accrual_start_col)+3)}{accrual_row}'] = accrual_differences[month]
                accrual_row += 1

            # Accumulate totals and add them to Excel
            total_original = sum(float(details.split()[0]) for details in original_monthly_accruals.values())
            total_bridge = sum(float(details.split()[0]) for details in bridge_monthly_accruals.values())
            total_difference = total_bridge - total_original

            # Add totals row
            ws[f'{accrual_start_col}{accrual_row}'] = "Total"
            ws[f'{chr(ord(accrual_start_col)+1)}{accrual_row}'] = f"{total_original:.2f} Hours"
            ws[f'{chr(ord(accrual_start_col)+2)}{accrual_row}'] = f"{total_bridge:.2f} Hours"
            ws[f'{chr(ord(accrual_start_col)+3)}{accrual_row}'] = f"{total_difference:.2f} Hours"

            # Auto size columns for better readability
            for col in ws.columns:
                max_length = max((len(str(cell.value)) for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2

            # Save the workbook
            file_name = f"{employee.employee_id} {employee.last_name}, {employee.first_name} Bridge In Service.xlsx"
            file_path = os.path.join(directory_path, file_name)
            wb.save(file_path)
            
            print(f"Data exported successfully to {file_path}")
            return True
        except Exception as e:
            print(f"Failed to export data to Excel: {e}")
            return False


class Email:

    @staticmethod
    def try_send_email(employee):

        try:
            pto_difference_line = ''
            if employee.pto_accrual_difference > 0:
                pto_difference_line = f"<strong>{employee.pto_accrual_difference:.2f}</strong> hours of PTO have been added to your accruals. You will see this reflected in your PTO bank within 1-2 paychecks. Please inform your payroll reporter these changes have been made."

            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.Subject = 'Bridge in Service'

            email_body = textwrap.dedent(f"""\
                <html>
                <head>
                    <style>
                        body {{ font-family: 'Century Gothic', sans-serif; }}
                    </style>
                </head>
                <body>
                    <p>Memorandum</p>
                    <p>To: {employee.first_name} {employee.last_name}<br>
                    From: University of Utah Hospitals and Clinics Benefits Department<br>
                    Date: {DateOperations.get_todays_date().strftime("%m/%d/%Y")}<br>
                    Subject: Service Date Adjustment</p>

                    <p>I am pleased to notify you that your application for a reinstatement of prior service has been approved. This process takes any 0.75 FTE (or above) benefited time and adds that to your most recent hire date.</p>

                    <p>Your new continuous “service date” is <strong>{employee.bridge_in_service_date.strftime("%m/%d/%Y")}</strong> rather than {employee.most_recent_start_date.strftime("%m/%d/%Y")}.</p>

                    <p>{pto_difference_line}</p>

                    <p>We are pleased that the Board of Trustees has approved this policy which recognizes the service of many valued employees such as you who rejoined the University following a break in service.</p>

                    <p>If you have any questions please feel free to contact the Human Resource Department at 801-581-6500.</p>

                    <p>Thank you,</p>

                    <p><strong>Benefits Department</strong><br>
                    Hospitals and Clinics</p>
              
                    <p><strong><span style="color: #BE0000;">Hospitals and Clinics Human Resources</span></strong><br>
                    525 E 100 S 1st Floor Suite 1810<br>
                    Salt Lake City UT 84102<br>
                    Ph 801.581.6500 | Fax 801.585.5144</p>
                </body>
                </html>
                """)

            mail.HTMLBody = email_body
            mail.To = f'u{employee.employee_id[1:]}@utah.edu'

            mail.Display()
            return True
        except Exception as e:
            print(f"Failed to send email: {e}")
        return False
